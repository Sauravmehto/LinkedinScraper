"""Validate company input Excel before running the pipeline.

Checks Official Website syntax, LinkedIn company URL syntax, and flags
verification / personal-profile / bad LinkedIn URLs.
Results are returned as structured dicts and also logged.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook

from gtm.linkedin_scraper.io_utils import (
    COMPANY_LINKEDIN_URL_HEADER,
    PROFILE_HEADER,
    WEBSITE_HEADER,
    find_column_index,
    get_header_row,
    resolve_worksheet,
)

LogFn = Callable[[str], None]

_BAD_LI_PATHS = re.compile(
    r"/(mycompany|verification|login|authwall|signIn|checkpoint|uas|lite|m|feed|"
    r"notifications|messaging|jobs|talent|learning|pulse|groups|school)/",
    re.I,
)
_PERSON_PROFILE_RE = re.compile(
    r"linkedin\.com/in/",
    re.I,
)
_COMPANY_SLUG_RE = re.compile(
    r"^https?://(?:[a-z]{2,3}\.)?linkedin\.com/company/[a-zA-Z0-9_%\-\.]+/?$",
    re.I,
)
_WEBSITE_BASIC_RE = re.compile(
    r"^https?://[^\s/$.?#][^\s]*$",
    re.I,
)


@dataclass
class RowValidation:
    row_num: int
    company_name: str
    website: str
    linkedin_url: str
    website_ok: bool = True
    website_warning: str = ""
    linkedin_ok: bool = True
    linkedin_warning: str = ""
    linkedin_was_corrected: bool = False
    linkedin_corrected_to: str = ""

    @property
    def has_warnings(self) -> bool:
        return bool(self.website_warning or self.linkedin_warning)

    def to_dict(self) -> dict:
        return {
            "row": self.row_num,
            "company": self.company_name,
            "website": self.website,
            "website_ok": self.website_ok,
            "website_warning": self.website_warning,
            "linkedin_url": self.linkedin_url,
            "linkedin_ok": self.linkedin_ok,
            "linkedin_warning": self.linkedin_warning,
            "linkedin_corrected_to": self.linkedin_corrected_to,
        }


def _normalize_url(url: str) -> str:
    url = (url or "").strip()
    if not url:
        return ""
    if not url.startswith(("http://", "https://")):
        url = "https://" + url
    return url


def _validate_website(website: str) -> tuple[bool, str]:
    raw = (website or "").strip()
    if not raw:
        return False, "empty"
    normalized = _normalize_url(raw)
    if not _WEBSITE_BASIC_RE.match(normalized):
        return False, f"invalid URL syntax: {raw!r}"
    return True, ""


def _validate_linkedin_company_url(url: str, *, company_name: str = "") -> tuple[bool, str, str]:
    """Return (is_ok, warning_message, corrected_url)."""
    raw = (url or "").strip()
    if not raw:
        return True, "", ""

    normalized = _normalize_url(raw)

    if _PERSON_PROFILE_RE.search(normalized):
        return False, "this looks like a personal /in/ URL, not a company page", ""

    if _BAD_LI_PATHS.search(normalized):
        clean = re.sub(r"/(mycompany|verification|login|authwall|signIn|checkpoint)[^\s]*$",
                       "", normalized, flags=re.I).rstrip("/")
        if _COMPANY_SLUG_RE.match(clean):
            return (
                False,
                f"URL contains a gated path ({url!r}); pipeline will try to use {clean!r}",
                clean,
            )
        return (
            False,
            f"URL contains a gated/invalid path that cannot be auto-corrected: {url!r}",
            "",
        )

    if not _COMPANY_SLUG_RE.match(normalized):
        return (
            False,
            f"does not match linkedin.com/company/{{slug}} format: {url!r}",
            "",
        )

    return True, "", ""


def validate_input_rows(
    rows: list[dict],
    *,
    log: LogFn | None = None,
) -> list[RowValidation]:
    _log = log or (lambda _m: None)
    results: list[RowValidation] = []

    for row in rows:
        rv = RowValidation(
            row_num=row.get("row_num", 0),
            company_name=row.get("company", ""),
            website=row.get("website", ""),
            linkedin_url=row.get("linkedin_url", ""),
        )

        ok_w, warn_w = _validate_website(rv.website)
        rv.website_ok = ok_w
        rv.website_warning = warn_w

        ok_li, warn_li, corrected = _validate_linkedin_company_url(
            rv.linkedin_url, company_name=rv.company_name
        )
        rv.linkedin_ok = ok_li
        rv.linkedin_warning = warn_li
        rv.linkedin_corrected_to = corrected
        rv.linkedin_was_corrected = bool(corrected)

        if rv.has_warnings:
            parts = [f"Row {rv.row_num} ({rv.company_name})"]
            if rv.website_warning:
                parts.append(f"website: {rv.website_warning}")
            if rv.linkedin_warning:
                parts.append(f"linkedin: {rv.linkedin_warning}")
            _log("  Input warning: " + " | ".join(parts))

        results.append(rv)

    return results


def validate_workbook(
    path: Path,
    *,
    sheet: str | None = None,
    log: LogFn | None = None,
) -> list[RowValidation]:
    _log = log or (lambda _m: None)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws, _name = resolve_worksheet(wb, sheet, log=_log)
    if ws is None:
        wb.close()
        return []

    headers = get_header_row(ws)
    website_col = find_column_index(headers, WEBSITE_HEADER)
    linkedin_col = find_column_index(headers, COMPANY_LINKEDIN_URL_HEADER)
    if linkedin_col is None:
        linkedin_col = find_column_index(headers, PROFILE_HEADER)

    raw_rows: list[dict] = []
    for row_idx in range(2, ws.max_row + 1):
        company = str(ws.cell(row=row_idx, column=1).value or "").strip()
        if not company:
            continue
        website = (
            str(ws.cell(row=row_idx, column=website_col + 1).value or "").strip()
            if website_col is not None
            else ""
        )
        linkedin = (
            str(ws.cell(row=row_idx, column=linkedin_col + 1).value or "").strip()
            if linkedin_col is not None
            else ""
        )
        raw_rows.append({
            "row_num": row_idx,
            "company": company,
            "website": website,
            "linkedin_url": linkedin,
        })

    wb.close()
    return validate_input_rows(raw_rows, log=_log)
