"""Validate Official Website URLs (HTTP status + final URL after redirects)."""

from __future__ import annotations

import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from typing import Optional

import httpx
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from gtm.linkedin_scraper.io_utils import (
    WEBSITE_FINAL_URL_HEADER,
    WEBSITE_HEADER,
    WEBSITE_STATUS_HEADER,
    LogFn,
    ensure_column,
    find_column_index,
    get_header_row,
    resolve_worksheet,
    write_column_values,
)

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)


def normalize_website(url: str) -> Optional[str]:
    url = (url or "").strip()
    if not url or url.lower() in ("n/a", "na", "-", "none"):
        return None
    if not url.startswith(("http://", "https://")):
        url = "https://" + url
    return url


def check_website_live(url: str, timeout: int) -> tuple[str, str]:
    normalized = normalize_website(url) or ""
    if not normalized:
        return "empty", ""

    headers = {
        "User-Agent": USER_AGENT,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
    }

    try:
        with httpx.Client(
            follow_redirects=True,
            timeout=timeout,
            headers=headers,
        ) as client:
            resp = client.get(normalized)
            final = str(resp.url)
            code = resp.status_code
    except httpx.TimeoutException:
        return "timeout", normalized
    except httpx.HTTPStatusError as exc:
        final = str(exc.response.url) if exc.response is not None else normalized
        code = exc.response.status_code if exc.response is not None else 0
    except httpx.HTTPError:
        return "timeout", normalized
    except OSError:
        return "timeout", normalized

    if code == 404:
        return "404", final
    if code in (403, 999):
        return "blocked", final
    if 200 <= code < 300:
        if final.rstrip("/") != normalized.rstrip("/"):
            return "OK (redirect)", final
        return "OK", final
    if code:
        return f"HTTP {code}", final
    return "error", final


def process_row(
    row_idx: int,
    website: str,
    timeout: int,
) -> tuple[int, str, str]:
    normalized = normalize_website(website)
    if not normalized:
        return row_idx, "empty", ""
    status, final = check_website_live(normalized, timeout)
    return row_idx, status, final


@dataclass
class WebsiteValidationStats:
    ok: int = 0
    redirect: int = 0
    timeout: int = 0
    blocked: int = 0
    notfound: int = 0
    other: int = 0


def run_validate_websites(
    wb: Workbook,
    *,
    sheet: Optional[str] = None,
    workers: int = 15,
    timeout: int = 15,
    limit: int = 0,
    log: LogFn = print,
) -> tuple[Optional[Worksheet], WebsiteValidationStats]:
    """Validate Official Website URLs in the workbook (in memory). Does not save."""
    ws, sheet_name = resolve_worksheet(
        wb, sheet, required_column=WEBSITE_HEADER, log=log
    )
    if ws is None:
        return None, WebsiteValidationStats()

    log(f"Validate websites [{sheet_name}]")

    headers = get_header_row(ws)
    website_col = find_column_index(headers, WEBSITE_HEADER)
    if website_col is None:
        log(f"Column not found: {WEBSITE_HEADER}")
        return None, WebsiteValidationStats()

    status_col, created_status = ensure_column(ws, headers, WEBSITE_STATUS_HEADER)
    if created_status:
        log(f"Added column: {WEBSITE_STATUS_HEADER}")

    final_col, created_final = ensure_column(ws, headers, WEBSITE_FINAL_URL_HEADER)
    if created_final:
        log(f"Added column: {WEBSITE_FINAL_URL_HEADER}")

    rows: list[tuple[int, str, str]] = []
    for row_idx in range(2, ws.max_row + 1):
        company = str(ws.cell(row=row_idx, column=1).value or "")
        website = str(ws.cell(row=row_idx, column=website_col + 1).value or "")
        rows.append((row_idx, company, website))

    if limit > 0:
        rows = rows[:limit]

    log(f"Rows to validate: {len(rows)} | workers={workers}")

    status_by_row: dict[int, str] = {}
    final_by_row: dict[int, str] = {}
    stats = WebsiteValidationStats()

    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {
            pool.submit(process_row, row_idx, website, timeout): (row_idx, company)
            for row_idx, company, website in rows
        }
        done = 0
        for fut in as_completed(futures):
            row_idx, company = futures[fut]
            done += 1
            try:
                r_idx, status, final = fut.result()
            except Exception as exc:
                log(f"[{done}/{len(rows)}] {company}: error {exc}")
                stats.other += 1
                continue

            status_by_row[r_idx] = status
            final_by_row[r_idx] = final

            if status == "OK":
                stats.ok += 1
            elif status == "OK (redirect)":
                stats.redirect += 1
            elif status == "timeout":
                stats.timeout += 1
            elif status == "blocked":
                stats.blocked += 1
            elif status == "404":
                stats.notfound += 1
            else:
                stats.other += 1

            log(f"[{done}/{len(rows)}] {company}: {status}")

    write_column_values(ws, status_col, status_by_row)
    write_column_values(ws, final_col, final_by_row)

    log(
        f"Website validation done. OK={stats.ok} redirect={stats.redirect} "
        f"404={stats.notfound} blocked={stats.blocked} timeout={stats.timeout}"
    )
    return ws, stats
