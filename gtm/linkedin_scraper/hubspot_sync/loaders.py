"""Load company and people rows from Excel outputs."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from gtm.linkedin_scraper.io_utils import (
    PROFILE_HEADER,
    WEBSITE_HEADER,
    find_column_index,
    get_header_row,
    resolve_worksheet,
)

from gtm.linkedin_scraper.people_discovery.contact_enrichment.people_excel import (
    read_people_workbook,
)

from .mapper import CompanyRow, company_row_from_excel


def _cell_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def load_companies_from_workbook(
    path: Path,
    *,
    sheet: str | None = None,
) -> list[CompanyRow]:
    wb = load_workbook(path, data_only=True)
    ws, _ = resolve_worksheet(wb, sheet)
    if ws is None:
        return []
    headers = get_header_row(ws)
    website_col = find_column_index(headers, WEBSITE_HEADER)
    profile_col = find_column_index(headers, PROFILE_HEADER)
    country_col = find_column_index(headers, "Country")
    headquarters_col = find_column_index(headers, "Headquarters")
    aum_col = find_column_index(headers, "Total AUM (USD)")
    asset_col = find_column_index(headers, "REIT Focus / Asset Type")
    phone_col = None
    for phone_header in (
        "HQ Phone",
        "Company Phone",
        "Phone",
        "Main Phone",
        "Headquarters Phone",
    ):
        phone_col = find_column_index(headers, phone_header)
        if phone_col is not None:
            break

    rows: list[CompanyRow] = []
    for row_idx in range(2, ws.max_row + 1):
        name = _cell_str(ws.cell(row=row_idx, column=1).value)
        website = (
            _cell_str(ws.cell(row=row_idx, column=website_col + 1).value)
            if website_col is not None
            else ""
        )
        profile = (
            _cell_str(ws.cell(row=row_idx, column=profile_col + 1).value)
            if profile_col is not None
            else ""
        )
        country = (
            _cell_str(ws.cell(row=row_idx, column=country_col + 1).value)
            if country_col is not None
            else ""
        )
        headquarters = (
            _cell_str(ws.cell(row=row_idx, column=headquarters_col + 1).value)
            if headquarters_col is not None
            else ""
        )
        aum = (
            _cell_str(ws.cell(row=row_idx, column=aum_col + 1).value)
            if aum_col is not None
            else ""
        )
        asset = (
            _cell_str(ws.cell(row=row_idx, column=asset_col + 1).value)
            if asset_col is not None
            else ""
        )
        hq_phone = (
            _cell_str(ws.cell(row=row_idx, column=phone_col + 1).value)
            if phone_col is not None
            else ""
        )
        row = company_row_from_excel(
            name=name,
            website=website,
            profile_url=profile,
            country=country,
            headquarters=headquarters,
            aum=aum,
            asset_type=asset,
            hq_phone=hq_phone,
        )
        if row:
            rows.append(row)
    return rows


def load_people_from_workbook(path: Path, *, sheet: str | None = None):
    return read_people_workbook(path, sheet=sheet)
