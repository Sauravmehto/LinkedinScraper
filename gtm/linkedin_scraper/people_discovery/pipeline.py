"""Decision-maker discovery pipeline orchestration."""

from __future__ import annotations

import json
from dataclasses import dataclass
from typing import Optional
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from gtm.linkedin_scraper.config import load_fallback_config, resolve_default_people_sources
from gtm.linkedin_scraper.io_utils import (
    OUTPUT_DIR,
    PROFILE_HEADER,
    WEBSITE_HEADER,
    find_column_index,
    get_header_row,
    resolve_worksheet,
)

from .anthropic_enrich import enrich_profile_hits
from .contact_enrichment import enrich_candidates
from .apollo_people import search_apollo_company
from .cache import load_cached_hits, save_cached_hits
from .candidate_extract import dedupe_candidates_advanced, dedupe_profile_hits
from .company_type import classify_company_type
from .quality import count_quality_profiles, is_quality_profile
from .company_match import company_mentioned_in_blob
from .query_builder import (
    build_all_role_queries,
    build_named_queries,
    build_priority_role_queries,
    build_tavily_people_queries,
    dedupe_queries,
)
from .role_map import expand_roles
from .scoring import confidence_from_score, score_candidate
from .search_sources import run_free_search_staged, run_search_sources, run_serper_fallback
from .team_pages import extract_team_linkedin_profiles, extract_team_people
from .linkedin_profile_enrich import enrich_linkedin_job_titles
from .title_clean import clean_candidates_job_titles
from .team_pages_playwright import shutdown_playwright
from .types import CompanyContext, DiscoveryStats, PersonCandidate, RawProfileHit, RoleTarget

MIN_CANDIDATE_SCORE = 55
MIN_CANDIDATE_SCORE_MAX = 55
DEFAULT_MIN_SCORE = 55


def _website_domain(website: str) -> str:
    raw = (website or "").strip()
    if not raw:
        return ""
    if not raw.startswith(("http://", "https://")):
        raw = "https://" + raw
    host = urlparse(raw).netloc.lower()
    if host.startswith("www."):
        host = host[4:]
    return host


def _company_match(hit: RawProfileHit, company_name: str, website: str = "") -> bool:
    if hit.source in ("apollo", "team_page", "playwright_team"):
        return True
    blob = f"{hit.title} {hit.snippet}"
    return company_mentioned_in_blob(blob, company_name, website=website)


def _role_relevant(
    hit: RawProfileHit,
    role_variations: list[str],
    company_name: str,
    website: str = "",
    *,
    max_coverage: bool = False,
) -> bool:
    if hit.source in ("apollo", "team_page", "playwright_team"):
        return True
    blob = f"{hit.title} {hit.snippet}"
    if not company_mentioned_in_blob(blob, company_name, website=website):
        return False
    if max_coverage and hit.source in ("bing", "ddg", "serper", "tavily"):
        return True
    return any(r.lower() in blob.lower() for r in role_variations)


def _source_rank(source: str) -> int:
    order = {
        "apollo": 0,
        "team_page": 1,
        "playwright_team": 1,
        "tavily": 2,
        "serper": 3,
        "bing": 4,
        "ddg": 5,
    }
    return order.get(source, 9)


def _needs_anthropic_cleanup(hits: list[RawProfileHit]) -> bool:
    messy_sources = {"bing", "ddg", "serper", "tavily"}
    for hit in hits:
        if hit.source in messy_sources:
            return True
        if hit.source in ("apollo", "team_page"):
            name = (hit.person_name or "").strip()
            if not name and hit.snippet:
                name = hit.snippet.split(" | ")[0].strip()
            if not is_quality_profile(hit, company_name=""):
                return True
    return False


def _hits_for_role(
    deduped: list[RawProfileHit],
    role: RoleTarget,
    company_name: str,
    website: str = "",
    *,
    max_coverage: bool = False,
) -> list[RawProfileHit]:
    return [
        h
        for h in deduped
        if _role_relevant(
            h, role.expanded_roles, company_name, website, max_coverage=max_coverage
        )
    ]


def _cap_company_candidates(
    candidates: list[PersonCandidate],
    *,
    max_people: int,
) -> list[PersonCandidate]:
    deduped = dedupe_candidates_advanced(candidates)
    deduped.sort(key=lambda c: (_source_rank(c.source), -c.score))

    if max_people <= 0:
        return deduped

    def _sort_key(c: PersonCandidate) -> tuple[int, int]:
        return (_source_rank(c.source), -c.score)

    by_role: dict[str, PersonCandidate] = {}
    for c in deduped:
        prev = by_role.get(c.role_target)
        if prev is None or _sort_key(c) < _sort_key(prev):
            by_role[c.role_target] = c

    diverse = sorted(by_role.values(), key=_sort_key)
    seen_urls = {c.linkedin_in_url.lower() for c in diverse}
    out = list(diverse[:max_people])

    if len(out) < max_people:
        for c in deduped:
            if len(out) >= max_people:
                break
            key = c.linkedin_in_url.lower()
            if key in seen_urls:
                continue
            seen_urls.add(key)
            out.append(c)

    return out[:max_people]


def _tavily_fallback_enabled(
    *,
    people_sources_explicit: bool,
    sources: tuple[str, ...],
    cfg_tavily_key: str | None,
) -> bool:
    if not cfg_tavily_key:
        return False
    if not people_sources_explicit:
        return True
    return "tavily" in sources


def _count_apollo_hits(hits: list[RawProfileHit]) -> int:
    return sum(1 for h in hits if h.source == "apollo")


def _gate_profile_count(
    hits: list[RawProfileHit],
    *,
    company_name: str,
    max_coverage: bool,
) -> int:
    """Profiles counted for paid-API skip gates."""
    deduped = dedupe_profile_hits(hits)
    if max_coverage:
        return len(deduped)
    return count_quality_profiles(deduped, company_name=company_name)


def _score_hit(
    hit: RawProfileHit,
    *,
    company_name: str,
    website: str,
    role_variations: list[str],
) -> int:
    person_name = _name_from_hit(hit)
    work_email = (hit.email or "").strip()
    score = score_candidate(
        company_name=company_name,
        role_variations=role_variations,
        title_text=hit.title,
        snippet_text=hit.snippet,
        person_name=person_name,
        work_email=work_email,
    )
    if hit.source == "apollo":
        score = min(100, score + 20)
    elif hit.source in ("team_page", "playwright_team"):
        score = min(100, score + 15)
    return score


def _best_score_for_hit(
    hit: RawProfileHit,
    *,
    company_name: str,
    website: str,
    role_targets: list[RoleTarget],
    max_coverage: bool,
) -> int:
    best = 0
    for role in role_targets:
        if not _role_relevant(
            hit,
            role.expanded_roles,
            company_name,
            website,
            max_coverage=max_coverage,
        ):
            continue
        if not _company_match(hit, company_name, website):
            continue
        best = max(
            best,
            _score_hit(
                hit,
                company_name=company_name,
                website=website,
                role_variations=role.expanded_roles,
            ),
        )
    return best


def _count_scored_hits_at_threshold(
    hits: list[RawProfileHit],
    *,
    company_name: str,
    website: str,
    role_targets: list[RoleTarget],
    min_score: int,
    max_coverage: bool,
) -> int:
    """Unique profiles whose best role score meets min_score (for API skip gates)."""
    if min_score <= 0:
        return _gate_profile_count(hits, company_name=company_name, max_coverage=max_coverage)
    count = 0
    for hit in dedupe_profile_hits(hits):
        if _best_score_for_hit(
            hit,
            company_name=company_name,
            website=website,
            role_targets=role_targets,
            max_coverage=max_coverage,
        ) >= min_score:
            count += 1
    return count


def _api_gate_count(
    hits: list[RawProfileHit],
    *,
    company_name: str,
    website: str,
    role_targets: list[RoleTarget],
    max_coverage: bool,
    min_score: int,
    use_scored_gates: bool,
) -> int:
    if use_scored_gates and min_score > 0:
        return _count_scored_hits_at_threshold(
            hits,
            company_name=company_name,
            website=website,
            role_targets=role_targets,
            min_score=min_score,
            max_coverage=max_coverage,
        )
    return _gate_profile_count(hits, company_name=company_name, max_coverage=max_coverage)


@dataclass
class DiscoverPeopleParams:
    sheet: Optional[str] = None
    company_type_column: str = "Company Type"
    max_candidates_per_role: int = 1
    max_people_per_company: int = 5
    apollo_sufficient_hits: int = 3
    min_people_before_serper: int = 3
    min_people_before_apollo: int = 3
    min_people_before_tavily: int = 2
    tavily_fallback_max_queries: int = 4
    serper_fallback_max_queries: int = 3
    max_queries_per_company: int = 8
    skip_ddg_if_bing_quality: int = 2
    people_sources: tuple[str, ...] | None = None
    timeout: int = 15
    workers: int = 8
    limit: int = 0
    enable_anthropic: bool = True
    anthropic_enrich_only: bool = False
    use_people_cache: bool = True
    refresh_cache: bool = False
    coverage_mode: str = "standard"
    enable_playwright_people: bool = False
    min_candidate_score: int = DEFAULT_MIN_SCORE
    force_tavily_fallback: bool = False
    use_scored_api_gates: bool = True
    enable_contact_enrichment: bool = False
    apollo_contact_delay: float = 0.5
    enrich_all_contacts: bool = False
    enable_linkedin_title_enrichment: bool = True
    linkedin_profile_delay: float = 2.0
    enable_title_clean: bool = True


def _name_from_profile_url(url: str) -> str:
    path = urlparse(url).path.strip("/")
    slug = path.split("/")[-1] if path else ""
    if not slug:
        return ""
    parts = [p for p in slug.replace("_", "-").split("-") if p]
    if not parts:
        return ""
    return " ".join(p.capitalize() for p in parts[:4])


def _name_from_hit(hit: RawProfileHit) -> str:
    if hit.person_name.strip():
        return hit.person_name.strip()
    snippet = hit.snippet or ""
    if " | " in snippet:
        maybe = snippet.split(" | ")[0].strip()
        if maybe and len(maybe.split()) <= 5:
            return maybe
    return _name_from_profile_url(hit.url)


def _all_role_variations(role_targets: list[RoleTarget]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for role in role_targets:
        for label in role.expanded_roles:
            key = label.lower()
            if key in seen:
                continue
            seen.add(key)
            out.append(label)
    return out


def _build_free_queries(
    ctx: CompanyContext,
    role_targets: list[RoleTarget],
    team_people: list[tuple[str, str]],
    *,
    max_queries: int,
    max_coverage: bool = False,
) -> list[str]:
    queries: list[str] = []
    if max_coverage:
        queries.extend(build_all_role_queries(ctx.name, role_targets))
    else:
        queries.extend(build_priority_role_queries(ctx.name, role_targets, max_roles=3))
    role_slice = role_targets if max_coverage else role_targets[:3]
    people_slice = team_people if max_coverage else team_people[:2]
    for role in role_slice:
        for person_name, _title in people_slice:
            queries.extend(build_named_queries(ctx.name, person_name, role))
    return dedupe_queries(queries, max_count=max_queries)


def _team_hits(
    profiles: list[tuple[str, str, str]],
) -> list[RawProfileHit]:
    hits: list[RawProfileHit] = []
    for url, name, title in profiles:
        hits.append(
            RawProfileHit(
                url=url,
                source="team_page",
                title=title,
                snippet=name,
                person_name=name,
                confidence_hint=0.9,
            )
        )
    return hits


def _to_company_contexts(
    ws: Worksheet,
    *,
    company_type_col_name: str,
    limit: int = 0,
) -> list[CompanyContext]:
    headers = get_header_row(ws)
    website_col = find_column_index(headers, WEBSITE_HEADER)
    profile_col = find_column_index(headers, PROFILE_HEADER)
    company_type_col = find_column_index(headers, company_type_col_name)
    if website_col is None:
        return []
    rows: list[CompanyContext] = []
    for row_idx in range(2, ws.max_row + 1):
        company = str(ws.cell(row=row_idx, column=1).value or "").strip()
        website = str(ws.cell(row=row_idx, column=website_col + 1).value or "").strip()
        company_linkedin = (
            str(ws.cell(row=row_idx, column=profile_col + 1).value or "").strip()
            if profile_col is not None
            else ""
        )
        if not company or not website:
            continue
        override = (
            str(ws.cell(row=row_idx, column=company_type_col + 1).value or "").strip()
            if company_type_col is not None
            else ""
        )
        domain = _website_domain(website)
        ctype = classify_company_type(
            company_name=company,
            website_text=domain,
            override=override,
        )
        rows.append(
            CompanyContext(
                row_idx=row_idx,
                name=company,
                website=website,
                company_linkedin=company_linkedin,
                company_type=ctype,
            )
        )
        if limit > 0 and len(rows) >= limit:
            break
    return rows


def run_people_discovery(
    wb: Workbook,
    *,
    params: DiscoverPeopleParams,
    log=print,
) -> tuple[list[PersonCandidate], DiscoveryStats]:
    ws, sheet_name = resolve_worksheet(wb, params.sheet, required_column=WEBSITE_HEADER, log=log)
    if ws is None:
        return [], DiscoveryStats()
    log(f"People discovery [{sheet_name}]")
    max_coverage = params.coverage_mode == "max"
    if max_coverage:
        log("Coverage mode: MAX (unlimited contacts, Playwright team pages, expanded roles)")
    cfg = load_fallback_config()
    sources_explicit = params.people_sources is not None
    sources = params.people_sources if params.people_sources is not None else resolve_default_people_sources()
    tavily_allowed = _tavily_fallback_enabled(
        people_sources_explicit=sources_explicit,
        sources=sources,
        cfg_tavily_key=cfg.tavily_api_key,
    ) or params.force_tavily_fallback
    log(
        f"People sources: {','.join(sources)}"
        + ("; tavily=fallback" if tavily_allowed else "")
        + ("; cache=on" if params.use_people_cache else "; cache=off")
    )
    if params.enable_anthropic and cfg.anthropic_api_key:
        log(f"Anthropic enrichment: on ({cfg.anthropic_model})")
    elif params.enable_anthropic:
        log("Anthropic enrichment: skipped (no ANTHROPIC_API_KEY)")
    if cfg.firecrawl_api_key:
        log("Firecrawl team pages: on (fallback when httpx returns thin HTML)")
    else:
        log("Firecrawl team pages: skipped (no FIRECRAWL_API_KEY)")
    if params.enable_playwright_people:
        log("Playwright team pages: on")
    else:
        log("Playwright team pages: off")
    log(
        f"Export min score: {params.min_candidate_score} "
        f"(paid API gates count profiles scoring >={params.min_candidate_score})"
    )

    if max_coverage or params.max_people_per_company <= 0:
        apollo_max_enrich = 25
    else:
        apollo_max_enrich = min(params.max_people_per_company + 2, 8)

    apollo_max_roles = 0 if max_coverage else 5

    contexts = _to_company_contexts(
        ws,
        company_type_col_name=params.company_type_column,
        limit=params.limit,
    )
    stats = DiscoveryStats(companies_checked=len(contexts))
    discovered: list[PersonCandidate] = []

    for idx, ctx in enumerate(contexts, 1):
        role_targets = expand_roles(ctx.company_type, max_coverage=max_coverage)
        all_variations = _all_role_variations(role_targets)
        unique_count = 0

        cache_status = "miss"
        raw_hits: list[RawProfileHit] = []
        apollo_status = "skipped"
        tavily_status = "skipped"
        serper_status = "skipped"
        serper_query_count = 0
        free_quality = 0

        if params.use_people_cache and not params.refresh_cache:
            cached = load_cached_hits(ctx.name, ctx.website)
            if cached is not None:
                raw_hits = cached
                cache_status = "hit"
                stats.cache_hits += 1

        if cache_status != "hit":
            team_profiles = extract_team_linkedin_profiles(
                website=ctx.website,
                role_variations=all_variations,
                timeout=float(params.timeout),
                firecrawl_api_key=cfg.firecrawl_api_key,
                enable_playwright=params.enable_playwright_people,
                max_coverage=max_coverage,
            )
            team_people = extract_team_people(
                website=ctx.website,
                role_variations=all_variations,
                timeout=float(params.timeout),
                firecrawl_api_key=cfg.firecrawl_api_key,
                enable_playwright=params.enable_playwright_people,
                max_coverage=max_coverage,
            )

            raw_hits.extend(_team_hits(team_profiles))

            query_cap = params.max_queries_per_company if params.max_queries_per_company > 0 else 0
            free_queries = _build_free_queries(
                ctx,
                role_targets,
                team_people,
                max_queries=query_cap,
                max_coverage=max_coverage,
            )
            use_bing = "bing" in sources
            use_ddg = "ddg" in sources
            if free_queries and (use_bing or use_ddg):
                raw_hits.extend(
                    run_free_search_staged(
                        queries=free_queries,
                        cfg=cfg,
                        timeout=float(params.timeout),
                        workers=params.workers,
                        use_bing=use_bing,
                        use_ddg=use_ddg,
                        skip_ddg_if_bing_quality_at_least=params.skip_ddg_if_bing_quality,
                        company_name=ctx.name,
                    )
                )

            after_free = dedupe_profile_hits(raw_hits)
            free_quality = _api_gate_count(
                after_free,
                company_name=ctx.name,
                website=ctx.website,
                role_targets=role_targets,
                max_coverage=max_coverage,
                min_score=params.min_candidate_score,
                use_scored_gates=params.use_scored_api_gates,
            )

            if "serper" not in sources:
                serper_status = "disabled"
            elif not cfg.serper_api_key:
                serper_status = "skipped_no_key"
            elif free_quality >= params.min_people_before_serper:
                serper_status = "skipped_sufficient"
            else:
                serper_cap = (
                    params.serper_fallback_max_queries
                    if params.serper_fallback_max_queries > 0
                    else 0
                )
                if max_coverage:
                    serper_queries = dedupe_queries(
                        build_all_role_queries(ctx.name, role_targets),
                        max_count=serper_cap,
                    )
                else:
                    serper_queries = dedupe_queries(
                        build_priority_role_queries(ctx.name, role_targets, max_roles=3),
                        max_count=serper_cap,
                    )
                serper_query_count = len(serper_queries)
                if serper_queries:
                    raw_hits.extend(
                        run_serper_fallback(
                            queries=serper_queries,
                            cfg=cfg,
                            timeout=float(params.timeout),
                            workers=params.workers,
                        )
                    )
                    serper_status = "used"
                else:
                    serper_status = "skipped_no_queries"

            after_serper = dedupe_profile_hits(raw_hits)
            quality_after_serper = _api_gate_count(
                after_serper,
                company_name=ctx.name,
                website=ctx.website,
                role_targets=role_targets,
                max_coverage=max_coverage,
                min_score=params.min_candidate_score,
                use_scored_gates=params.use_scored_api_gates,
            )

            if (
                "apollo" in sources
                and cfg.apollo_api_key
                and quality_after_serper < params.min_people_before_apollo
            ):
                raw_hits.extend(
                    search_apollo_company(
                        company_name=ctx.name,
                        website=ctx.website,
                        role_targets=role_targets,
                        api_key=cfg.apollo_api_key,
                        timeout=float(params.timeout),
                        max_enrich=apollo_max_enrich,
                        max_roles=apollo_max_roles if apollo_max_roles > 0 else 99,
                    )
                )
                apollo_status = "used"
            else:
                apollo_status = "skipped_free_ok"

            deduped_mid = dedupe_profile_hits(raw_hits)
            quality_total = _api_gate_count(
                deduped_mid,
                company_name=ctx.name,
                website=ctx.website,
                role_targets=role_targets,
                max_coverage=max_coverage,
                min_score=params.min_candidate_score,
                use_scored_gates=params.use_scored_api_gates,
            )

            if tavily_allowed and quality_total < params.min_people_before_tavily:
                tavily_cap = (
                    params.tavily_fallback_max_queries
                    if params.tavily_fallback_max_queries > 0
                    else 0
                )
                tavily_queries = dedupe_queries(
                    build_tavily_people_queries(
                        ctx.name,
                        role_targets,
                        max_roles=0 if max_coverage else 5,
                    ),
                    max_count=tavily_cap,
                )
                if tavily_queries:
                    raw_hits.extend(
                        run_search_sources(
                            queries=tavily_queries,
                            enabled_sources=("tavily",),
                            cfg=cfg,
                            timeout=float(params.timeout),
                            workers=params.workers,
                        )
                    )
                    tavily_status = "used"
            elif not tavily_allowed:
                tavily_status = "disabled"
            else:
                tavily_status = "skipped_sufficient"

            if params.use_people_cache:
                save_cached_hits(
                    ctx.name,
                    ctx.website,
                    dedupe_profile_hits(raw_hits),
                    meta={
                        "free_quality": free_quality,
                        "serper_status": serper_status,
                        "apollo_status": apollo_status,
                        "tavily_status": tavily_status,
                    },
                )

        deduped = dedupe_profile_hits(raw_hits)
        apollo_count = _count_apollo_hits(deduped)
        quality_total = _api_gate_count(
            deduped,
            company_name=ctx.name,
            website=ctx.website,
            role_targets=role_targets,
            max_coverage=max_coverage,
            min_score=params.min_candidate_score,
            use_scored_gates=params.use_scored_api_gates,
        )
        unique_count = len(deduped)

        skip_anthropic = (
            quality_total >= params.apollo_sufficient_hits
            and not _needs_anthropic_cleanup(deduped)
        )
        anthropic_ran = False
        if (
            params.enable_anthropic
            and cfg.anthropic_api_key
            and deduped
            and not skip_anthropic
            and cache_status != "hit"
        ):
            deduped = enrich_profile_hits(
                company_name=ctx.name,
                role_targets=role_targets,
                hits=deduped,
                api_key=cfg.anthropic_api_key,
                model=cfg.anthropic_model or "claude-sonnet-4-5-20250929",
                timeout=max(float(params.timeout), 30.0),
                enrich_only=params.anthropic_enrich_only or max_coverage,
            )
            deduped = dedupe_profile_hits(deduped)
            anthropic_ran = True

        if not params.enable_anthropic or not cfg.anthropic_api_key:
            anthropic_log = "disabled"
        elif skip_anthropic or cache_status == "hit":
            anthropic_log = "skipped"
        elif anthropic_ran:
            anthropic_log = "used"
        else:
            anthropic_log = "skipped"

        company_candidates: list[PersonCandidate] = []
        min_score = params.min_candidate_score

        for role in role_targets:
            role_hits = _hits_for_role(
                deduped, role, ctx.name, ctx.website, max_coverage=max_coverage
            )
            ranked: list[PersonCandidate] = []
            for hit in role_hits:
                if not _company_match(hit, ctx.name, ctx.website):
                    continue
                person_name = _name_from_hit(hit)
                work_email = (hit.email or "").strip()
                score = _score_hit(
                    hit,
                    company_name=ctx.name,
                    website=ctx.website,
                    role_variations=role.expanded_roles,
                )
                if score < min_score:
                    continue
                email_status = "from_apollo" if work_email and hit.source == "apollo" else ""
                email_confidence = "from_apollo" if work_email and hit.source == "apollo" else ""
                direct_dial = (hit.direct_dial or "").strip()
                hq_phone = (hit.hq_phone or "").strip()
                phone_source = (hit.phone_source or "").strip()
                if (direct_dial or hq_phone) and not phone_source:
                    phone_source = "apollo" if hit.source == "apollo" else hit.source
                ranked.append(
                    PersonCandidate(
                        company_name=ctx.name,
                        company_type=ctx.company_type,
                        company_linkedin=ctx.company_linkedin,
                        company_website=ctx.website,
                        role_target=role.primary_role,
                        person_name=person_name,
                        person_title=hit.title or role.primary_role,
                        linkedin_in_url=hit.url,
                        source=hit.source,
                        snippet=hit.snippet,
                        score=score,
                        confidence=confidence_from_score(score),
                        notes="",
                        work_email=work_email,
                        email_status=email_status,
                        email_confidence=email_confidence,
                        direct_dial=direct_dial,
                        hq_phone=hq_phone,
                        phone_source=phone_source,
                    )
                )

            ranked.sort(key=lambda c: (_source_rank(c.source), -c.score))
            if params.max_candidates_per_role <= 0:
                company_candidates.extend(ranked)
            else:
                company_candidates.extend(ranked[: params.max_candidates_per_role])

        company_candidates = _cap_company_candidates(
            company_candidates,
            max_people=params.max_people_per_company,
        )

        source_counts: dict[str, int] = {}
        for c in company_candidates:
            stats.bump_source(c.source)
            source_counts[c.source] = source_counts.get(c.source, 0) + 1

        if company_candidates:
            stats.companies_with_candidates += 1
        stats.candidates_total += len(company_candidates)
        discovered.extend(company_candidates)
        src_summary = ",".join(f"{k}={v}" for k, v in sorted(source_counts.items())) or "none"
        serper_log = serper_status
        if serper_status == "used" and serper_query_count:
            serper_log = f"used({serper_query_count})"
        log(
            f"[{idx}/{len(contexts)}] {ctx.name}: "
            f"{len(company_candidates)} candidate(s) | unique_hits={unique_count} | "
            f"scored>={params.min_candidate_score}={quality_total} | "
            f"cache={cache_status} | free_q={free_quality} | "
            f"serper={serper_log} | apollo={apollo_status}({apollo_count}) | "
            f"tavily={tavily_status} | anthropic={anthropic_log} | "
            f"{src_summary}"
        )

    if params.enable_playwright_people:
        shutdown_playwright()

    if params.enable_contact_enrichment:
        cfg = load_fallback_config()
        websites_by_company = {ctx.name: ctx.website for ctx in contexts if ctx.name and ctx.website}
        discovered, enrich_stats = enrich_candidates(
            discovered,
            api_key=cfg.apollo_api_key,
            timeout=float(params.timeout),
            only_missing=not params.enrich_all_contacts,
            apollo_delay=params.apollo_contact_delay,
            websites_by_company=websites_by_company,
            log=log,
        )
        log(
            f"Contact columns: work_email={enrich_stats.with_work_email} "
            f"personal_email={enrich_stats.with_personal_email} "
            f"direct_dial={enrich_stats.with_direct_dial} "
            f"hq_phone={enrich_stats.with_hq_phone}"
        )

    if params.enable_linkedin_title_enrichment:
        cfg = load_fallback_config()
        discovered, title_stats = enrich_linkedin_job_titles(
            discovered,
            api_key=cfg.anthropic_api_key,
            enable_anthropic=params.enable_anthropic,
            profile_delay=params.linkedin_profile_delay,
            timeout=float(params.timeout),
            log=log,
        )
        log(
            f"Job titles: fetched={title_stats.profiles_fetched} "
            f"updated={title_stats.titles_updated} "
            f"verified={title_stats.anthropic_verified}"
        )

    if params.enable_title_clean:
        cfg = load_fallback_config()
        discovered, clean_stats = clean_candidates_job_titles(
            discovered,
            api_key=cfg.anthropic_api_key,
            enable_anthropic=params.enable_anthropic,
            timeout=float(params.timeout),
            log=log,
        )
        log(
            f"Title clean: updated={clean_stats.cleaned} "
            f"unknown={clean_stats.unknown}"
        )
        if clean_stats.results:
            json_path = OUTPUT_DIR / "title_clean_results.json"
            json_path.parent.mkdir(parents=True, exist_ok=True)
            json_path.write_text(
                json.dumps(clean_stats.results, indent=2, ensure_ascii=False),
                encoding="utf-8",
            )
            log(f"Title clean JSON: {json_path}")

    return discovered, stats
