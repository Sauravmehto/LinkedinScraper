"""Read/write decision_makers Excel with contact columns."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from gtm.linkedin_scraper.io_utils import WEBSITE_HEADER, find_column_index, get_header_row, resolve_worksheet

from ..types import CompanyType, Confidence, PersonCandidate


def _resolve_people_ws(wb, sheet: str | None):
    if sheet and sheet in wb.sheetnames:
        return wb[sheet]
    if "Decision Makers" in wb.sheetnames:
        return wb["Decision Makers"]
    return wb.active


PEOPLE_EXPORT_COLUMNS: tuple[str, ...] = (
    "company_name",
    "company_type",
    "company_linkedin",
    "company_website",
    "person_name",
    "person_title",
    "person_linkedin",
    "work_email",
    "email_confidence",
    "email_status",
    "direct_dial",
    "hq_phone",
    "ir_email",
    "ir_phone",
    "phone_source",
    "phone_status",
    "city",
    "state",
    "country",
    "source",
    "score",
    "confidence",
    "role_target",
    "notes",
)

_LEGACY_ALIASES: dict[str, tuple[str, ...]] = {
    "person_linkedin": ("person_linkedin", "linkedin_in_url"),
    "company_website": ("company_website", "official_website"),
}


def _cell_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _row_dict(headers: list[str], row_values: tuple) -> dict[str, str]:
    out: dict[str, str] = {}
    for idx, header in enumerate(headers):
        key = (header or "").strip().lower()
        if not key:
            continue
        val = row_values[idx] if idx < len(row_values) else ""
        out[key] = _cell_str(val)
    return out


def _get_field(row: dict[str, str], field: str) -> str:
    keys = _LEGACY_ALIASES.get(field, (field,))
    for key in keys:
        if key in row and row[key]:
            return row[key]
    return row.get(field, "")


def _parse_confidence(raw: str) -> Confidence:
    upper = (raw or "").strip().upper()
    if upper in ("HIGH", "MEDIUM", "LOW"):
        return upper  # type: ignore[return-value]
    return "LOW"


def _parse_company_type(raw: str) -> CompanyType:
    upper = (raw or "").strip().upper()
    allowed = ("PE", "REIT", "HEDGE_FUND", "VC", "FAMILY_OFFICE", "UNKNOWN")
    if upper in allowed:
        return upper  # type: ignore[return-value]
    return "UNKNOWN"


def candidate_from_row(row: dict[str, str]) -> PersonCandidate | None:
    linkedin = _get_field(row, "person_linkedin")
    company = _get_field(row, "company_name")
    if not linkedin or not company:
        return None
    score_raw = _get_field(row, "score") or "0"
    try:
        score = int(float(score_raw))
    except ValueError:
        score = 0
    email_conf = _get_field(row, "email_confidence")
    if not email_conf and _get_field(row, "work_email"):
        src = _get_field(row, "email_status")
        email_conf = "from_apollo" if src == "from_apollo" else "unknown"
    return PersonCandidate(
        company_name=company,
        company_type=_parse_company_type(_get_field(row, "company_type")),
        company_linkedin=_get_field(row, "company_linkedin"),
        company_website=_get_field(row, "company_website"),
        role_target=_get_field(row, "role_target"),
        person_name=_get_field(row, "person_name"),
        person_title=_get_field(row, "person_title"),
        linkedin_in_url=linkedin,
        source=_get_field(row, "source") or "import",
        snippet="",
        score=score,
        confidence=_parse_confidence(_get_field(row, "confidence")),
        notes=_get_field(row, "notes"),
        work_email=_get_field(row, "work_email"),
        email_status=_get_field(row, "email_status"),
        email_confidence=email_conf,
        direct_dial=_get_field(row, "direct_dial"),
        hq_phone=_get_field(row, "hq_phone"),
        ir_email=_get_field(row, "ir_email"),
        ir_phone=_get_field(row, "ir_phone"),
        phone_source=_get_field(row, "phone_source"),
        phone_status=_get_field(row, "phone_status"),
        city=_get_field(row, "city"),
        state=_get_field(row, "state"),
        country=_get_field(row, "country"),
    )


def candidate_to_row(c: PersonCandidate) -> list:
    return [
        c.company_name,
        c.company_type,
        c.company_linkedin,
        c.company_website,
        c.person_name,
        c.person_title,
        c.linkedin_in_url,
        c.work_email,
        c.email_confidence,
        c.email_status,
        c.direct_dial,
        c.hq_phone,
        c.ir_email,
        c.ir_phone,
        c.phone_source,
        c.phone_status,
        c.city,
        c.state,
        c.country,
        c.source,
        c.score,
        c.confidence,
        c.role_target,
        c.notes,
    ]


def read_people_workbook(
    path: Path,
    *,
    sheet: str | None = None,
) -> list[PersonCandidate]:
    wb = load_workbook(path, data_only=True)
    ws = _resolve_people_ws(wb, sheet)
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = [_cell_str(v).lower() for v in header_row]
    candidates: list[PersonCandidate] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
        row_dict = _row_dict(headers, row)
        cand = candidate_from_row(row_dict)
        if cand:
            candidates.append(cand)
    return candidates


def write_people_workbook(path: Path, candidates: list[PersonCandidate]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Decision Makers"
    ws.append(list(PEOPLE_EXPORT_COLUMNS))
    for c in candidates:
        ws.append(candidate_to_row(c))
    try:
        wb.save(path)
    except PermissionError:
        alt = path.with_name(f"{path.stem}_new{path.suffix}")
        wb.save(alt)
        raise PermissionError(
            f"Could not write {path} (file may be open in Excel). Saved to {alt} instead."
        ) from None


_COMPANY_PHONE_HEADERS = (
    "hq phone",
    "company phone",
    "phone",
    "main phone",
    "headquarters phone",
)


def load_company_hq_phones(
    path: Path,
    *,
    sheet: str | None = None,
    company_column: int = 1,
) -> dict[str, str]:
    """Map company name (column A) -> HQ/main phone from optional columns in company workbook."""
    if not path.exists():
        return {}
    wb = load_workbook(path, data_only=True)
    ws, _sheet_name = resolve_worksheet(wb, sheet)
    if ws is None:
        return {}
    headers = get_header_row(ws)
    phone_col = None
    for name in _COMPANY_PHONE_HEADERS:
        phone_col = find_column_index(headers, name)
        if phone_col is not None:
            break
    if phone_col is None:
        return {}
    out: dict[str, str] = {}
    for row_idx in range(2, ws.max_row + 1):
        company = _cell_str(ws.cell(row=row_idx, column=company_column).value)
        phone = _cell_str(ws.cell(row=row_idx, column=phone_col + 1).value)
        if company and phone:
            out[company] = phone
    return out


def load_company_websites(
    path: Path,
    *,
    sheet: str | None = None,
    company_column: int = 1,
) -> dict[str, str]:
    """Map company name (column A) -> Official Website from company result workbook."""
    if not path.exists():
        return {}
    wb = load_workbook(path, data_only=True)
    ws, _sheet_name = resolve_worksheet(wb, sheet)
    if ws is None:
        ws = _resolve_people_ws(wb, sheet)
    headers = get_header_row(ws)
    website_col = find_column_index(headers, WEBSITE_HEADER)
    if website_col is None:
        return {}
    out: dict[str, str] = {}
    for row_idx in range(2, ws.max_row + 1):
        company = _cell_str(ws.cell(row=row_idx, column=company_column).value)
        website = _cell_str(ws.cell(row=row_idx, column=website_col + 1).value)
        if company and website:
            out[company] = website
    return out
