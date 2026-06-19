"""HubSpot Data.xlsx template headers and row mapping."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from gtm.linkedin_scraper.hubspot_sync.client import domain_from_website
from gtm.linkedin_scraper.hubspot_sync.mapper import CompanyRow, split_person_name
from gtm.linkedin_scraper.io_utils import DATA_DIR
from gtm.linkedin_scraper.people_discovery.candidate_extract import normalize_profile_url
from gtm.linkedin_scraper.people_discovery.types import PersonCandidate

from .template_map import _clean_last_name, _city_from_headquarters, resolve_company_for_candidate

HUBSPOT_DATA_TEMPLATE = DATA_DIR / "Hubspot Data.xlsx"


def load_template_headers(template_path: Path) -> list[str]:
    wb = load_workbook(template_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        return [str(c.value or "").strip() for c in ws[1] if c.value is not None]
    finally:
        wb.close()


def _state_region(candidate: PersonCandidate, company: CompanyRow | None) -> str:
    country = (candidate.country or "").strip()
    if country:
        return country
    if company and company.country:
        return company.country.strip()
    state = (candidate.state or "").strip()
    if state:
        return state
    if company and company.headquarters and "," in company.headquarters:
        return company.headquarters.split(",", 1)[1].strip()
    return ""


def _city(candidate: PersonCandidate, company: CompanyRow | None) -> str:
    city = (candidate.city or "").strip()
    if city:
        return city
    if company and company.headquarters:
        return _city_from_headquarters(company.headquarters)
    return ""


def candidate_to_hubspot_row(
    candidate: PersonCandidate,
    company: CompanyRow | None,
) -> dict[str, Any]:
    """Map merged candidate + company to Hubspot Data.xlsx column names."""
    first, last = split_person_name(candidate.person_name)
    person_li = normalize_profile_url(candidate.linkedin_in_url)
    last = _clean_last_name(last, person_li)
    title = (candidate.person_title or candidate.role_target or "").strip()
    email = (
        (candidate.work_email or candidate.personal_email or "").strip().lower()
    )
    company_li = normalize_profile_url(candidate.company_linkedin)
    website = (candidate.company_website or "").strip()
    if not website and company:
        website = (company.website or "").strip()
    if not company_li and company:
        company_li = normalize_profile_url(company.company_linkedin)
    mobile = (candidate.direct_dial or "").strip()
    company_phone = (candidate.hq_phone or "").strip()
    if not company_phone and company:
        company_phone = (company.hq_phone or "").strip()
    domain = domain_from_website(website) if website else ""
    industry = (company.asset_type or "").strip() if company else ""

    return {
        "First Name": first,
        "Last Name": last,
        "Job Title": title,
        "Email": email,
        "Company Name": (candidate.company_name or "").strip(),
        "Linkedin account": person_li,
        "Mobile Phone Number": mobile,
        "Company Phone Number": company_phone,
        "Company Domain Name": domain,
        "Industry": industry,
        "City": _city(candidate, company),
        "State/Region": _state_region(candidate, company),
        "Number of Employees": "",
        "LinkedIn Company Page": company_li,
    }


def candidate_to_source_payload(
    candidate: PersonCandidate,
    company: CompanyRow | None,
) -> dict[str, Any]:
    """Wide payload for Claude (includes scores and company firmographics)."""
    row = candidate_to_hubspot_row(candidate, company)
    row["_score"] = candidate.score
    row["_confidence"] = candidate.confidence
    row["_role_target"] = candidate.role_target
    row["_source"] = candidate.source
    row["_company_website"] = (candidate.company_website or "").strip() or (
        (company.website or "").strip() if company else ""
    )
    if company:
        row["_aum"] = company.aum
        row["_asset_focus"] = company.asset_type
        row["_company_country"] = company.country
        row["_company_headquarters"] = company.headquarters
    return row


def build_hubspot_rows(
    candidates: list[PersonCandidate],
    companies_by_name: dict[str, CompanyRow],
    *,
    indexes=None,
) -> list[dict[str, Any]]:
    from .company_lookup import build_company_indexes

    idx = indexes or build_company_indexes(list(companies_by_name.values()))
    out: list[dict[str, Any]] = []
    for candidate in candidates:
        company = resolve_company_for_candidate(
            candidate, companies_by_name, indexes=idx
        )
        out.append(candidate_to_hubspot_row(candidate, company))
    return out


def build_source_payloads(
    candidates: list[PersonCandidate],
    companies_by_name: dict[str, CompanyRow],
) -> list[dict[str, Any]]:
    from .company_lookup import build_company_indexes

    idx = build_company_indexes(list(companies_by_name.values()))
    out: list[dict[str, Any]] = []
    for candidate in candidates:
        company = resolve_company_for_candidate(
            candidate, companies_by_name, indexes=idx
        )
        out.append(candidate_to_source_payload(candidate, company))
    return out


def row_dict_to_values(headers: list[str], row: dict[str, Any]) -> list[Any]:
    return [row.get(h, "") for h in headers]


def normalize_row_dict(row: dict[str, Any], headers: list[str]) -> dict[str, Any]:
    """Keep only template columns; coerce values to strings."""
    out: dict[str, Any] = {}
    for h in headers:
        val = row.get(h, "")
        if val is None:
            val = ""
        out[h] = str(val).strip() if val != "" else ""
    return out
