"""Write final_report.xlsx from HubSpot template + merged contacts."""

from __future__ import annotations

import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from gtm.linkedin_scraper.hubspot_sync.mapper import CompanyRow
from gtm.linkedin_scraper.people_discovery.types import PersonCandidate

from .company_lookup import build_company_indexes
from .hubspot_data_template import load_template_headers, row_dict_to_values
from .template_map import ReportDefaults, build_row_values, resolve_company_for_candidate


def write_hubspot_data_report(
    template_path: Path,
    output_path: Path,
    headers: list[str],
    row_dicts: list[dict[str, Any]],
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, output_path)

    try:
        wb = load_workbook(output_path)
    except PermissionError as exc:
        raise PermissionError(
            f"Cannot open {output_path}. Close Excel and retry."
        ) from exc

    ws = wb.active
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for row in row_dicts:
        ws.append(row_dict_to_values(headers, row))

    try:
        wb.save(output_path)
    except PermissionError as exc:
        raise PermissionError(
            f"Cannot save {output_path}. Close Excel and retry."
        ) from exc
    finally:
        wb.close()


def write_final_report(
    template_path: Path,
    output_path: Path,
    candidates: list[PersonCandidate],
    companies_by_name: dict[str, CompanyRow],
    *,
    defaults: ReportDefaults | None = None,
    dry_run: bool = False,
) -> None:
    """Legacy writer: map candidates to template headers via template_map."""
    if dry_run:
        return

    headers = load_template_headers(template_path)
    company_indexes = build_company_indexes(list(companies_by_name.values()))
    row_dicts: list[dict[str, Any]] = []
    for candidate in candidates:
        company = resolve_company_for_candidate(
            candidate, companies_by_name, indexes=company_indexes
        )
        values = build_row_values(
            headers,
            candidate,
            company=company,
            defaults=defaults,
        )
        row_dicts.append(dict(zip(headers, values)))

    write_hubspot_data_report(template_path, output_path, headers, row_dicts)
