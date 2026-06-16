"""Shared openpyxl helpers for GTM Excel workflows."""

from __future__ import annotations

import re
import shutil
from pathlib import Path
from typing import Callable, Iterator, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

REPO_ROOT = Path(__file__).resolve().parents[2]
DATA_DIR = REPO_ROOT / "data"
OUTPUT_DIR = REPO_ROOT / "output"
DEFAULT_XLSX = DATA_DIR / "TargetReachout.xlsx"
DEFAULT_SAMPLE_XLSX = DATA_DIR / "Sample_file.xlsx"
SHEET_NAME = "REIT Asset Managers"
WEBSITE_HEADER = "Official Website"
PROFILE_HEADER = "Profile URL"
SCRAPE_METHOD_HEADER = "Scrape Method"
WEBSITE_STATUS_HEADER = "Website Status"
WEBSITE_FINAL_URL_HEADER = "Website Final URL"
URL_VALID_HEADER = "URL Valid (syntax)"
URL_STATUS_HEADER = "URL Status (live)"

PROFILE_COLUMN_ALIASES = (
    "profile url",
    "linkedin url",
    "linkedin profile",
    "linkedin company url",
    "linkedin link",
)

LINKEDIN_CELL_RE = re.compile(
    r"https?://(?:[a-z]{2,3}\.)?linkedin\.com/(?:company|in|school)/[a-zA-Z0-9_%\-\.]+/?",
    re.IGNORECASE,
)

LogFn = Callable[[str], None]


def default_output_path(input_path: Path) -> Path:
    return OUTPUT_DIR / f"{input_path.stem}_enriched.xlsx"


def prepare_workbook(
    input_path: Path,
    output_path: Path,
    *,
    dry_run: bool,
) -> Workbook:
    """Load workbook for processing. Never modifies *input_path* when dry-run or output differs."""
    if dry_run:
        return load_workbook(input_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if input_path.resolve() != output_path.resolve():
        shutil.copy2(input_path, output_path)
    return load_workbook(output_path)


def save_workbook(wb: Workbook, output_path: Path, *, dry_run: bool) -> None:
    if dry_run:
        return
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def find_column_index(headers: list, name: str) -> Optional[int]:
    for i, h in enumerate(headers):
        if h and str(h).strip().lower() == name.lower():
            return i
    return None


def get_header_row(ws: Worksheet) -> list:
    return [cell.value for cell in ws[1]]


def sheet_has_column(ws: Worksheet, column_name: str) -> bool:
    return find_column_index(get_header_row(ws), column_name) is not None


def sheet_has_website_column(ws: Worksheet) -> bool:
    return sheet_has_column(ws, WEBSITE_HEADER)


def find_profile_column(headers: list) -> Optional[int]:
    for alias in PROFILE_COLUMN_ALIASES:
        col = find_column_index(headers, alias)
        if col is not None:
            return col
    for i, h in enumerate(headers):
        if not h:
            continue
        label = str(h).strip().lower()
        if "linkedin" in label and any(
            token in label for token in ("url", "profile", "link", "company")
        ):
            return i
    return None


def resolve_validation_worksheet(
    wb: Workbook,
    sheet_arg: Optional[str],
    log: LogFn = lambda _msg: None,
) -> tuple[Optional[Worksheet], Optional[str]]:
    """Pick a sheet with Profile URL, or fall back to Official Website (e.g. Sheet1)."""
    ws, name = resolve_worksheet(
        wb,
        sheet_arg,
        required_column=PROFILE_HEADER,
        preferred_sheet=SHEET_NAME,
        log=log,
    )
    if ws is not None:
        return ws, name
    return resolve_worksheet(
        wb,
        sheet_arg,
        required_column=WEBSITE_HEADER,
        preferred_sheet=SHEET_NAME,
        log=log,
    )


def resolve_worksheet(
    wb: Workbook,
    sheet_arg: Optional[str],
    *,
    required_column: str = WEBSITE_HEADER,
    preferred_sheet: str = SHEET_NAME,
    log: LogFn = lambda _msg: None,
) -> tuple[Optional[Worksheet], Optional[str]]:
    if sheet_arg:
        if sheet_arg not in wb.sheetnames:
            log(f"Sheet not found: {sheet_arg}")
            log(f"Available sheets: {', '.join(wb.sheetnames)}")
            return None, None
        ws = wb[sheet_arg]
        if not sheet_has_column(ws, required_column):
            log(f"Column not found on sheet '{sheet_arg}': {required_column}")
            return None, None
        return ws, sheet_arg

    if preferred_sheet in wb.sheetnames:
        ws = wb[preferred_sheet]
        if sheet_has_column(ws, required_column):
            return ws, preferred_sheet

    for name in wb.sheetnames:
        ws = wb[name]
        if sheet_has_column(ws, required_column):
            if name != preferred_sheet:
                log(f"Using sheet '{name}' (auto-detected via {required_column})")
            return ws, name

    if wb.active is not None and sheet_has_column(wb.active, required_column):
        log(f"Warning: using active sheet '{wb.active.title}'")
        return wb.active, wb.active.title

    missing: list[str] = []
    for name in wb.sheetnames:
        if not sheet_has_column(wb[name], required_column):
            missing.append(name)
    log(f"No sheet with column '{required_column}' found.")
    log(f"Sheets: {', '.join(wb.sheetnames)}")
    if missing:
        log(f"Sheets missing {required_column}: {', '.join(missing)}")
    return None, None


def ensure_column(ws: Worksheet, headers: list, name: str) -> tuple[int, bool]:
    """Return 0-based column index; True if the column was newly created."""
    col = find_column_index(headers, name)
    if col is not None:
        return col, False
    col = len(headers)
    ws.cell(row=1, column=col + 1, value=name)
    headers.append(name)
    return col, True


def iter_data_row_indices(ws: Worksheet, start_row: int = 2) -> Iterator[int]:
    for row_idx in range(start_row, ws.max_row + 1):
        yield row_idx


def build_scrape_tasks(
    ws: Worksheet,
    website_col: int,
    profile_col: int,
    *,
    skip_filled: bool,
    company_col: int = 0,
) -> list[tuple[int, str, str, Optional[object]]]:
    tasks: list[tuple[int, str, str, Optional[object]]] = []
    for row_idx in iter_data_row_indices(ws):
        company = ws.cell(row=row_idx, column=company_col + 1).value or ""
        website = ws.cell(row=row_idx, column=website_col + 1).value
        existing = ws.cell(row=row_idx, column=profile_col + 1).value
        if website or (skip_filled and existing):
            tasks.append((row_idx, str(company), str(website or ""), existing))
    return tasks


def write_column_values(
    ws: Worksheet,
    col_idx: int,
    values_by_row: dict[int, object],
    *,
    only_if_value: bool = False,
) -> None:
    for row_idx, value in values_by_row.items():
        if only_if_value and not value:
            continue
        ws.cell(row=row_idx, column=col_idx + 1, value=value or "")


def extract_linkedin_from_cell(value: object) -> Optional[str]:
    if not value:
        return None
    match = LINKEDIN_CELL_RE.search(str(value))
    return match.group(0) if match else None


def collect_profile_urls(
    ws: Worksheet,
    profile_col: int,
) -> list[tuple[int, str, str]]:
    """Return (row_idx, company_name, profile_url) for non-empty Profile URL cells."""
    rows: list[tuple[int, str, str]] = []
    for row_idx in iter_data_row_indices(ws):
        company = ws.cell(row=row_idx, column=1).value or ""
        url = extract_linkedin_from_cell(
            ws.cell(row=row_idx, column=profile_col + 1).value
        )
        if url:
            rows.append((row_idx, str(company), url))
    return rows


def collect_urls_for_validation(
    ws: Worksheet,
    profile_col: int,
) -> list[tuple[int, str, str]]:
    """Profile URL column first; if empty, scan any cell in the row for LinkedIn links."""
    rows = collect_profile_urls(ws, profile_col)
    if rows:
        return rows

    rows = []
    for row_idx in iter_data_row_indices(ws):
        company = ws.cell(row=row_idx, column=1).value or ""
        url: Optional[str] = None
        for col in range(1, ws.max_column + 1):
            found = extract_linkedin_from_cell(ws.cell(row=row_idx, column=col).value)
            if found:
                url = found
                break
        if url:
            rows.append((row_idx, str(company), url))
    return rows
